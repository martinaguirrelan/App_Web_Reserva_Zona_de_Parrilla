from uuid import UUID
from io import BytesIO
from datetime import date
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from ..database import get_db
from ..models.reservation import Reservation
from ..models.zone import Zone
from ..models.payment import Payment
from ..schemas.zone import ZoneCreate, ZoneUpdate, ZoneResponse
from ..schemas.reservation import ReservationResponse, ReservationStatusUpdate
from .auth import get_current_admin

router = APIRouter(prefix="/admin", tags=["admin"])

VALID_STATES = ("pendiente_pago", "en_revision", "confirmada", "rechazada", "cancelada", "pendiente_devolucion")


# ── Reservas ──────────────────────────────────────────────────────────────────

@router.get("/reservations", response_model=list[ReservationResponse])
def list_reservations(
    estado: str | None = None,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    q = db.query(Reservation)
    if estado:
        q = q.filter(Reservation.estado == estado)
    return q.order_by(Reservation.created_at.desc()).all()


@router.patch("/reservations/{reservation_id}/estado")
def update_estado(
    reservation_id: UUID,
    data: ReservationStatusUpdate,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    if data.estado not in VALID_STATES:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Opciones: {VALID_STATES}")

    reserva = db.query(Reservation).filter(Reservation.id == reservation_id).first()
    if not reserva:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")

    reserva.estado = data.estado
    if data.notas_admin and reserva.pagos:
        reserva.pagos[-1].notas_admin = data.notas_admin
    if data.monto_garantia_dev is not None:
        reserva.monto_garantia_dev = data.monto_garantia_dev
    if data.monto_limpieza is not None:
        reserva.monto_limpieza = data.monto_limpieza

    db.commit()
    return {"mensaje": "Estado actualizado", "estado": data.estado}


@router.get("/reservations/{reservation_id}/payment")
def get_payment_info(
    reservation_id: UUID,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    """Devuelve la info del comprobante de pago para ver en el panel admin."""
    reserva = db.query(Reservation).filter(Reservation.id == reservation_id).first()
    if not reserva:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    if not reserva.pagos:
        raise HTTPException(status_code=404, detail="Sin comprobante cargado.")
    pago = reserva.pagos[-1]
    return {
        "archivo_url": pago.archivo_url,
        "archivo_nombre": pago.archivo_nombre,
        "notas_admin": pago.notas_admin,
        "created_at": pago.created_at,
    }


# ── Stats ─────────────────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    """Dashboard: totales, ingresos y distribución por estado y zona."""
    total = db.query(func.count(Reservation.id)).scalar() or 0

    por_estado = (
        db.query(Reservation.estado, func.count(Reservation.id))
        .group_by(Reservation.estado)
        .all()
    )

    ingresos_confirmados = (
        db.query(func.sum(Reservation.monto_total))
        .filter(Reservation.estado == "confirmada")
        .scalar()
        or Decimal("0")
    )

    ingresos_totales = (
        db.query(func.sum(Reservation.monto_total))
        .filter(Reservation.estado.in_(("pendiente_pago", "en_revision", "confirmada")))
        .scalar()
        or Decimal("0")
    )

    por_zona = (
        db.query(Zone.nombre, func.count(Reservation.id))
        .join(Reservation, Reservation.zone_id == Zone.id)
        .filter(Reservation.estado.in_(("pendiente_pago", "en_revision", "confirmada")))
        .group_by(Zone.nombre)
        .all()
    )

    # Reservas de los últimos 30 días
    from datetime import datetime, timedelta
    hace_30 = datetime.utcnow() - timedelta(days=30)
    recientes = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.created_at >= hace_30)
        .scalar()
        or 0
    )

    return {
        "total_reservas": total,
        "reservas_recientes_30d": recientes,
        "ingresos_confirmados": float(ingresos_confirmados),
        "ingresos_proyectados": float(ingresos_totales),
        "por_estado": {e: c for e, c in por_estado},
        "por_zona": {z: c for z, c in por_zona},
    }


# ── Exportar Excel ────────────────────────────────────────────────────────────

@router.get("/reservations/export")
def export_excel(
    estado: str | None = None,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    """Genera y descarga un archivo Excel con las reservas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    q = db.query(Reservation)
    if estado:
        q = q.filter(Reservation.estado == estado)
    reservas = q.order_by(Reservation.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas"

    header_fill = PatternFill("solid", fgColor="D97706")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Código", "Residente", "Departamento", "Email",
        "Zona", "Fecha", "Hora inicio", "Hora fin",
        "Monto (S/.)", "Estado", "Notas", "Fecha creación",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ESTADO_LABELS = {
        "pendiente_pago": "Pendiente de Pago",
        "en_revision": "En Revisión",
        "confirmada": "Confirmada",
        "rechazada": "Rechazada",
        "cancelada": "Cancelada",
    }

    for r in reservas:
        ws.append([
            r.codigo,
            r.usuario.nombre if r.usuario else "",
            r.usuario.departamento if r.usuario else "",
            r.usuario.email if r.usuario else "",
            r.zona.nombre if r.zona else "",
            r.fecha.strftime("%d/%m/%Y") if r.fecha else "",
            r.hora_inicio.strftime("%H:%M") if r.hora_inicio else "",
            r.hora_fin.strftime("%H:%M") if r.hora_fin else "",
            float(r.monto_total),
            ESTADO_LABELS.get(r.estado, r.estado),
            r.notas or "",
            r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reservas{'_' + estado if estado else ''}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Zonas ─────────────────────────────────────────────────────────────────────

@router.get("/zones", response_model=list[ZoneResponse])
def list_all_zones(
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    return db.query(Zone).order_by(Zone.created_at).all()


@router.post("/zones", response_model=ZoneResponse, status_code=201)
def create_zone(
    data: ZoneCreate,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    zone = Zone(**data.model_dump())
    db.add(zone)
    db.commit()
    db.refresh(zone)
    return zone


@router.put("/zones/{zone_id}", response_model=ZoneResponse)
def update_zone(
    zone_id: UUID,
    data: ZoneUpdate,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    zone = db.query(Zone).filter(Zone.id == zone_id).first()
    if not zone:
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(zone, field, value)
    db.commit()
    db.refresh(zone)
    return zone


@router.delete("/zones/{zone_id}", status_code=204)
def deactivate_zone(
    zone_id: UUID,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    zone = db.query(Zone).filter(Zone.id == zone_id).first()
    if not zone:
        raise HTTPException(status_code=404, detail="Zona no encontrada")
    zone.activa = False
    db.commit()


# ── Seed ──────────────────────────────────────────────────────────────────────

@router.post("/seed", status_code=201)
def seed_zones(
    db: Session = Depends(get_db),
    _: str = Depends(get_current_admin),
):
    if db.query(Zone).count() > 0:
        raise HTTPException(status_code=400, detail="Ya existen zonas configuradas.")

    defaults = [
        Zone(
            nombre="Zona Parrilla Simple",
            descripcion="Parrilla cubierta para hasta 10 personas. Incluye carbón y utensilios básicos.",
            precio_base=150,
        ),
        Zone(
            nombre="Zona Parrilla + SUM Bar",
            descripcion="Parrilla completa con acceso al SUM y barra. Ideal para eventos grandes.",
            precio_base=280,
        ),
    ]
    db.add_all(defaults)
    db.commit()
    return {"mensaje": f"{len(defaults)} zonas creadas correctamente."}
